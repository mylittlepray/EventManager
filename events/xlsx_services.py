import openpyxl

from io import BytesIO 
from openpyxl.writer.excel import save_workbook

from django.http import HttpResponse
from django.contrib.gis.geos import Point
from django.db import transaction

from venues.models import Venue
from .models import Event, EventStatus

def parse_coordinates(coord_str):
    """
    Парсит строку вида 'longitude, latitude' (например '37.61, 55.75') в Point.
    """
    try:
        lon, lat = map(float, coord_str.replace(";", ",").split(","))
        return Point(lon, lat, srid=4326)
    except (ValueError, AttributeError):
        return None

def export_events_to_xlsx(queryset):
    """
    Генерирует XLSX-файл из QuerySet событий.
    Возвращает HttpResponse с файлом.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Events"

    # Заголовки (согласно ТЗ)
    headers = [
        "Дата публикации", 
        "Дата начала", 
        "Дата завершения", 
        "Место проведения", 
        "Рейтинг"
    ]
    ws.append(headers)

    for event in queryset:
        ws.append([
            event.publish_at.strftime("%Y-%m-%d %H:%M") if event.publish_at else "",
            event.start_at.strftime("%Y-%m-%d %H:%M"),
            event.end_at.strftime("%Y-%m-%d %H:%M"),
            event.venue.name,
            event.rating
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    content = buffer.getvalue()

    response = HttpResponse(
        content=content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="events_export.xlsx"'
    return response


def import_events_from_xlsx(file_obj, user):
    """
    Читает XLSX файл и создает события.
    Возвращает статистику (создано, ошибок).
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active
    
    created_count = 0
    errors = []

    # Начинаем со 2-й строки
    rows = ws.iter_rows(min_row=2, values_only=True)
    
    for i, row in enumerate(rows, start=2):
        # ВАЖНО: atomic на каждую строку отдельно!
        # Если эта строка упадет - откатится только она.
        # Если пройдёт - сохранится сразу.
        try:
            with transaction.atomic():
                if not row or not row[0]: # Пропускаем пустые
                    continue

                # ... (твой код распаковки row) ...
                title = row[0]
                description = row[1] or ""
                publish_at = row[2] 
                start_at = row[3]
                end_at = row[4]
                venue_name = row[5]
                coords_str = str(row[6])
                rating = row[7] or 0
                
                # Парсинг Venue
                point = parse_coordinates(coords_str)
                if not point:
                    venue = Venue.objects.filter(name=venue_name).first()
                    if not venue:
                        raise ValueError(f"Venue '{venue_name}' not found and no coords")
                else:
                    venue, _ = Venue.objects.get_or_create(
                        name=venue_name,
                        defaults={"location": point}
                    )

                # Создание Event
                Event.objects.create(
                    title=title,
                    description=description,
                    publish_at=publish_at,
                    start_at=start_at,
                    end_at=end_at,
                    venue=venue,
                    rating=rating,
                    author=user,
                    status=EventStatus.DRAFT
                )
                created_count += 1

        except Exception as e:
            # Ошибка в конкретной строке ловится тут, транзакция этой строки откатывается,
            # но цикл идет дальше к следующей строке.
            errors.append(f"Row {i}: {str(e)}")

    return {"created": created_count, "errors": errors}
